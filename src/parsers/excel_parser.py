"""Excel file parser for fund data ingestion."""

import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class ExcelParser:
    """Parse Excel files and extract fund data rows."""

    def __init__(self, engine: str = "openpyxl"):
        """
        Initialize Excel parser.

        Args:
            engine: openpyxl or xlrd (openpyxl is default for .xlsx)
        """
        self.engine = engine

    def parse(self, file_path: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """
        Parse Excel file and return list of row dictionaries.

        Args:
            file_path: Path to Excel file (.xlsx)
            sheet_name: Name or index of sheet (default: 0, first sheet)

        Returns:
            List of dictionaries, one per row (header names as keys)

        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If file cannot be parsed
        """
        path = Path(file_path)

        if not path.exists():
            logger.error(f"Excel file not found: {file_path}")
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        try:
            # Use sheet_name parameter (can be int or str)
            sheet_param = sheet_name if sheet_name is not None else 0

            df = pd.read_excel(file_path, sheet_name=sheet_param, engine=self.engine)

            logger.debug(f"Parsed Excel sheet '{sheet_param}': {file_path}")

            # Convert Excel dates (datetime objects) to ISO-8601 strings
            for col in df.columns:
                if pd.api.types.is_datetime64_any_dtype(df[col]):
                    df[col] = df[col].dt.strftime("%Y-%m-%d")
                    logger.debug(f"Converted datetime column '{col}' to ISO-8601 strings")

            # Convert to list of dicts
            rows = df.to_dict(orient="records")
            logger.info(f"Successfully parsed {len(rows)} rows from {file_path}")
            return rows

        except FileNotFoundError as e:
            logger.error(f"Sheet not found in {file_path}: {e}")
            raise ValueError(f"Sheet not found: {e}")
        except pd.errors.ParserError as e:
            logger.error(f"Excel parsing error in {file_path}: {e}")
            raise ValueError(f"Excel parsing error: {e}")
        except Exception as e:
            logger.error(f"Unexpected error parsing Excel {file_path}: {e}")
            raise ValueError(f"Error parsing Excel: {e}")

    def extract_rows_with_line_numbers(
        self, file_path: str, sheet_name: Optional[str] = None
    ) -> List[Tuple[int, Dict[str, Any]]]:
        """
        Parse Excel file and return list of (row_number, row_dict) tuples.

        Row numbers are 1-based (excluding header).

        Args:
            file_path: Path to Excel file
            sheet_name: Name or index of sheet (default: 0)

        Returns:
            List of (row_number, row_dict) tuples

        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If file cannot be parsed
        """
        rows = self.parse(file_path, sheet_name)
        # Row numbers are 1-based (row 1 is first data row after header)
        return [(i + 1, row) for i, row in enumerate(rows)]

    def list_sheets(self, file_path: str) -> List[str]:
        """
        List all sheet names in Excel file.

        Args:
            file_path: Path to Excel file

        Returns:
            List of sheet names

        Raises:
            FileNotFoundError: If file doesn't exist
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            sheets = workbook.sheetnames
            logger.debug(f"Found {len(sheets)} sheets in {file_path}: {sheets}")
            return sheets
        except Exception as e:
            logger.error(f"Error listing sheets in {file_path}: {e}")
            raise ValueError(f"Error reading Excel file: {e}")

    def validate_structure(
        self, file_path: str, sheet_name: Optional[str] = None, required_columns: List[str] = None
    ) -> bool:
        """
        Validate Excel structure (header columns).

        Args:
            file_path: Path to Excel file
            sheet_name: Name or index of sheet (default: 0)
            required_columns: List of required column names (optional)

        Returns:
            True if valid, False otherwise

        Raises:
            ValueError: If required columns are missing
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        try:
            sheet_param = sheet_name if sheet_name is not None else 0
            df = pd.read_excel(file_path, sheet_name=sheet_param, nrows=0, engine=self.engine)
            columns = df.columns.tolist()

            if required_columns:
                missing = set(required_columns) - set(columns)
                if missing:
                    raise ValueError(f"Excel missing required columns: {missing}")

            logger.debug(f"Excel structure valid: {columns}")
            return True

        except Exception as e:
            logger.error(f"Excel structure validation failed: {e}")
            raise ValueError(f"Excel validation error: {e}")
